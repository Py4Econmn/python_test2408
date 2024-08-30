from openpyxl import load_workbook
import pandas as pd
import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border


def copy_style(source_cell, target_cell):
    # Copy font
    target_cell.font = Font(
        name=source_cell.font.name,
        bold=source_cell.font.bold,
        italic=source_cell.font.italic,
        vertAlign=source_cell.font.vertAlign,
        underline=source_cell.font.underline,
        strike=source_cell.font.strike,
        color=source_cell.font.color
    )
    # Copy fill
    target_cell.fill = PatternFill(
        start_color=source_cell.fill.start_color,
        end_color=source_cell.fill.end_color,
        fill_type=source_cell.fill.fill_type
    )
    # Copy alignment
    target_cell.alignment = Alignment(
        horizontal=source_cell.alignment.horizontal,
        vertical=source_cell.alignment.vertical,
        text_rotation=source_cell.alignment.text_rotation,
        wrap_text=source_cell.alignment.wrap_text,
        shrink_to_fit=source_cell.alignment.shrink_to_fit,
        indent=source_cell.alignment.indent
    )
    # Copy border
    target_cell.border = Border(
        left=source_cell.border.left,
        right=source_cell.border.right,
        top=source_cell.border.top,
        bottom=source_cell.border.bottom,
        diagonal=source_cell.border.diagonal,
        diagonal_direction=source_cell.border.diagonal_direction,
        outline=source_cell.border.outline
    )

current_month = 6 # current_month = datetime.datetime.now().month - 2


# Load the workbook
# wb = load_workbook('data/bop.xlsx')

# # Select the active worksheet (or a specific sheet)
# sheet = wb.active

# wb.sheetnames

# # cell values
# sheet['B5'].value
# sheet.cell(5,2).value

# sheet['B5'].value = 9999

# # Save the workbook
# wb.save('data/bop_bom.xlsx')

# df_bop_bom = pd.read_excel('data/bop.xlsx',skiprows=1)

# columns = df_bop_bom.columns
# df_bop_bom['cum'] = df_bop_bom[columns[-current_month:]].sum(axis=1)
# df_bom = df_bop_bom.iloc[:,-2:]

# wb = load_workbook('data/bop_nso.xlsx')
# ws = wb['11.2']
# ws['A1'].value 
# ws['U6'].value 


# for i in range(6, len(df_bom)):
#     ws['V' + str(i)].value = df_bom.iloc[i-6,0]
#     ws['W' + str(i)].value = df_bom.iloc[i-6,1]

# for row in ws.iter_rows(min_row=6, max_row=8, min_col=ws.max_column-2, max_col=ws.max_column):
#     for cell in row:
#         print(cell.value)

# wb.save('data/bop_nso2.xlsx')


wb = load_workbook('data/bop_nso.xlsx')
ws = wb['11.1']
ws.insert_rows(15)
ws.delete_rows(13)
ws.cell(14,2).value = 'I-VII'
# ws.cell(14,3).value = sum(ws.cell(row=row, column=3).value for row in range(7, 13 + 1))
# ws.cell(14,4).value = sum(ws.cell(row=row, column=4).value for row in range(7, 13 + 1))
# ws.cell(14,5).value = sum(ws.cell(row=row, column=5).value for row in range(7, 13 + 1))
# ws.cell(14,6).value = sum(ws.cell(row=row, column=6).value for row in range(7, 13 + 1))
# ws.cell(14,7).value = sum(ws.cell(row=row, column=7).value for row in range(7, 13 + 1))

start_row = 7
end_row = 13
columns = [3, 4, 5, 6, 7]

number_format = '0.00'

# Calculate the sum for each column and set values
for col in columns:
    total_sum = sum(ws.cell(row=row, column=col).value for row in range(start_row, end_row + 1))
    rounded_sum = round(total_sum, 2)  # Round to two decimal places
    cell = ws.cell(row=14, column=col)
    cell.value = rounded_sum
    cell.number_format = number_format  # Apply number formatting

    cell.font = Font(underline='single') 
    
    # Copy formatting from cell above (row 13) to the new row (row 14)
    source_cell = ws.cell(row=13, column=col)
    copy_style(source_cell, cell)

    cell.number_format = number_format

wb.save('data/bop_nso2.xlsx')
